import openpyxl
import requests as rq
from bs4 import BeautifulSoup


def get_page(page):
    answer = []
    url = f'https://prozhito.org/api/notes/&offset={page * 25}'
    resp = rq.get(url)
    data = resp.json()
    data = data['data']
    items = data['notes']
    for item in items:
        author = item['person']['firstName'] + " " + item['person']['lastName'] + " " + item['person']['thirdName']
        data = item['date']
        data = data.split('-')
        data = reversed(data)
        data = '-'.join(data)
        age = item['age']
        text = BeautifulSoup(item['text'], "html.parser").text
        answer.append((author, data, age, text))
    return answer


def main():
    items = []
    page = 0
    while (True):
        tempory = get_page(page)
        if (len(tempory)):
            page += 1
            items += tempory
            print(tempory)
        else:
            break
    wb = openpyxl.Workbook()
    ws = wb.active
    for index, item in enumerate(items):
        author, data, age, text = item
        index += 1
        ws[f'A{index}'] = author
        ws[f'B{index}'] = data
        ws[f'C{index}'] = age
        ws[f'D{index}'] = text
    wb.save("prozhito.xlsx")


if __name__ == '__main__':
    main()
